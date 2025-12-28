import asyncio
import pandas as pd
import json
import os
import base64
from datetime import datetime
from bilibili_api import user, channel_series, video, Credential, sync

# --- 新增：加密库 ---
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

# 引入 openpyxl 用于样式调整
from openpyxl.styles import Border, Side, Alignment, Font

# --- 1. 安全 Cookie 管理模块 ---
COOKIE_FILE = "cookies.bin"  # 改为二进制文件

def _derive_key(password: str, salt: bytes) -> bytes:
    """根据密码和盐值生成密钥"""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
    )
    return base64.urlsafe_b64encode(kdf.derive(password.encode()))

def save_cookies_encrypted(sessdata, bili_jct, buvid3):
    """加密保存 Cookies"""
    data = json.dumps({"SESSDATA": sessdata, "BILI_JCT": bili_jct, "BUVID3": buvid3})
    
    while True:
        pwd = input("🔐 请设置一个读取密码 (用于加密本地文件): ").strip()
        if pwd:
            break
        print("密码不能为空！")

    # 1. 生成随机盐值
    salt = os.urandom(16)
    # 2. 生成密钥
    key = _derive_key(pwd, salt)
    f = Fernet(key)
    # 3. 加密数据
    encrypted_data = f.encrypt(data.encode())
    
    # 4. 保存：前16字节是盐值，后面是加密内容
    with open(COOKIE_FILE, "wb") as f_out:
        f_out.write(salt + encrypted_data)
        
    print(f"💡 凭据已加密并保存至 {COOKIE_FILE}")

def load_cookies_encrypted():
    """解密读取 Cookies"""
    if not os.path.exists(COOKIE_FILE):
        return None

    print(f"📂 发现加密的凭据文件: {COOKIE_FILE}")
    pwd = input("🔑 请输入密码以解密: ").strip()

    try:
        with open(COOKIE_FILE, "rb") as f_in:
            file_content = f_in.read()
        
        # 提取盐值 (前16字节) 和 密文
        salt = file_content[:16]
        encrypted_data = file_content[16:]
        
        # 还原密钥
        key = _derive_key(pwd, salt)
        f = Fernet(key)
        
        # 解密
        decrypted_data = f.decrypt(encrypted_data)
        return json.loads(decrypted_data.decode())

    except InvalidToken:
        print("❌ 密码错误！无法解密。")
        return None
    except Exception as e:
        print(f"❌ 文件损坏或读取错误: {e}")
        return None

async def get_credential():
    # 尝试加载本地加密凭据
    cookies = load_cookies_encrypted()
    
    if cookies:
        print("🔓 解密成功，正在验证有效性...")
        c = Credential(
            sessdata=cookies['SESSDATA'], 
            bili_jct=cookies['BILI_JCT'], 
            buvid3=cookies['BUVID3']
        )
        try:
            if await c.check_valid(): 
                print("✅ 登录验证通过！")
                return c
            else:
                print("⚠️ 本地凭据已过期。")
        except: 
            print("⚠️ 网络连接失败或凭据无效。")
    
    # 如果没有文件、密码错误或凭据过期，重新输入
    print("\n" + "="*40)
    print("🆕 需要重新登录 (获取方式：浏览器 F12 -> Application -> Cookies)")
    s = input("SESSDATA: ").strip()
    j = input("bili_jct: ").strip()
    b = input("buvid3: ").strip()
    
    # 验证新输入的凭据
    c = Credential(sessdata=s, bili_jct=j, buvid3=b)
    if await c.check_valid():
        save_cookies_encrypted(s, j, b) # 保存时会要求设置密码
        return c
    else:
        print("❌ 输入的 Cookies 无效，请检查。")
        return None

# --- 2. 核心抓取逻辑 (保持功能不变) ---

async def fetch_live_series_data(credential, uid, sid, year):
    print(f"\n📺 [1/3] 正在扫描直播合集 (ID: {sid})...")
    cs = channel_series.ChannelSeries(uid=uid, id_=sid, type_=channel_series.ChannelSeriesType.SERIES, credential=credential)
    
    live_data = []
    live_bvids = set()
    page = 1
    series_name = "未命名合集"
    
    try:
        meta = await cs.get_meta()
        series_name = meta.get('meta', {}).get('name', '未命名合集')
        print(f"   合集名称: {series_name}")

        while True:
            res = await cs.get_videos(pn=page, ps=30)
            archives = res.get('archives', [])
            if not archives: break
            
            for v in archives:
                pub_ts = v.get('pubdate', v.get('ctime'))
                dt = datetime.fromtimestamp(pub_ts)
                if year and dt.year != year: continue
                
                bvid = v['bvid']
                live_bvids.add(bvid)
                
                stat = v.get('stat', {})
                view = stat.get('view', 0)
                like = stat.get('like', 0)
                coin = stat.get('coin', 0)
                fav = stat.get('favorite', 0)
                triple = like + coin + fav
                
                live_data.append({
                    "日期": dt.strftime("%Y-%m-%d"),
                    "标题": v['title'],
                    "类型": "直播回放",
                    "时长(小时)": round(v.get('duration', 0) / 3600, 2),
                    "播放": view,
                    "点赞": like,
                    "投币": coin,
                    "收藏": fav,
                    "三连": triple,
                    "BV号": bvid,
                    "链接": f"https://www.bilibili.com/video/{bvid}"
                })
            
            if page * 30 >= res.get('page', {}).get('count', 0): break
            page += 1
            print(f"\r   已获取 {len(live_data)} 条回放数据...", end="")
            await asyncio.sleep(0.2)
            
    except Exception as e:
        print(f"\n❌ 获取合集失败: {e}")
    
    print(f"\n   ✅ 直播回放扫描完毕，共 {len(live_data)} 条。")
    return live_data, live_bvids, series_name

async def fetch_user_uploads(credential, uid, year, exclude_bvids):
    print(f"\n📹 [2/3] 正在扫描主页投稿 (排除直播合集)...")
    u = user.User(uid=uid, credential=credential)
    video_data = []
    page = 1
    is_finish = False
    
    while not is_finish:
        res = await u.get_videos(pn=page, ps=30)
        vlist = res.get('list', {}).get('vlist', [])
        if not vlist: break

        for v in vlist:
            dt = datetime.fromtimestamp(v['created'])
            if year:
                if dt.year > year: continue
                if dt.year < year: 
                    is_finish = True
                    break
            
            bvid = v['bvid']
            if bvid in exclude_bvids: continue 
            
            print(f"\r   正在获取详情: {v['title'][:15]}...", end="")
            try:
                v_obj = video.Video(bvid=bvid, credential=credential)
                info = await v_obj.get_info()
                stat = info['stat']
                
                video_data.append({
                    "日期": dt.strftime("%Y-%m-%d"),
                    "标题": v['title'],
                    "类型": "普通投稿",
                    "时长(小时)": round(info['duration'] / 3600, 2),
                    "播放": stat['view'],
                    "点赞": stat['like'],
                    "投币": stat['coin'],
                    "收藏": stat['favorite'],
                    "三连": stat['like'] + stat['coin'] + stat['favorite'],
                    "BV号": bvid,
                    "链接": f"https://www.bilibili.com/video/{bvid}"
                })
                await asyncio.sleep(0.3)
            except Exception as e:
                print(f" (跳过: {e})", end="")
        
        if is_finish: break
        page += 1
        
    print(f"\n   ✅ 普通投稿扫描完毕，共 {len(video_data)} 条。")
    return video_data

# --- 3. 样式调整函数 ---
def style_excel(writer, sheet_name, title_text):
    worksheet = writer.sheets[sheet_name]
    
    # 插入年份标题行
    worksheet.insert_rows(1)
    worksheet['A1'] = title_text
    worksheet.merge_cells('A1:C1')
    
    # 样式定义
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    title_font = Font(name='微软雅黑', size=12, bold=False)  
    header_font = Font(name='微软雅黑', size=11, bold=True)   
    
    # 应用顶部标题样式
    cell_title = worksheet['A1']
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    cell_title.font = title_font
    cell_title.border = thin_border
    worksheet['B1'].border = thin_border
    worksheet['C1'].border = thin_border

    # 遍历所有行
    for row in worksheet.iter_rows(min_row=2):
        cell_a = row[0]
        cell_b = row[1]
        
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

        # 合并分隔行 (A-C)
        val = str(cell_a.value) if cell_a.value else ""
        if ("---" in val or "榜" in val) and cell_b.value == "":
            worksheet.merge_cells(start_row=cell_a.row, start_column=1, end_row=cell_a.row, end_column=3)
            cell_a.font = header_font
            cell_a.alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头行
        if cell_a.row == 2:
            for cell in row:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列宽
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 45

# --- 4. 辅助函数 ---
def get_top_n(data_list, key, n=5):
    sorted_list = sorted(data_list, key=lambda x: x[key], reverse=True)
    return sorted_list[:n]

async def generate_report():
    print("="*50)
    print("   Bilibili UP主年度数据综合分析工具 (加密版)")
    print("="*50)

    credential = await get_credential()
    if not credential: return

    try:
        uid = int(input("1. 输入UP主 UID: ").strip())
        sid = int(input("2. 输入直播回放合集 ID (sid): ").strip())
        year_in = input("3. 统计年份 (例如 2025，回车统计全部): ").strip()
        target_year = int(year_in) if year_in.isdigit() else None
    except ValueError:
        print("❌ 输入错误。")
        return

    # 获取数据
    u = user.User(uid=uid, credential=credential)
    try:
        u_info = await u.get_user_info()
        nickname = u_info['name']
        u_rel = await u.get_relation_info()
        fans = u_rel['follower']
    except Exception as e:
        print(f"❌ 获取UP主信息失败，可能是凭据失效: {e}")
        return

    live_list, live_bvids, series_name = await fetch_live_series_data(credential, uid, sid, target_year)
    video_list = await fetch_user_uploads(credential, uid, target_year, live_bvids)
    all_data = live_list + video_list
    
    if not all_data:
        print("❌ 无数据。")
        return

    # 计算
    df_all = pd.DataFrame(all_data)
    total_likes = df_all['点赞'].sum()
    total_coins = df_all['投币'].sum()
    total_favs = df_all['收藏'].sum()
    
    live_hours = sum(i['时长(小时)'] for i in live_list)
    live_count = len(live_list)
    live_likes = sum(i['点赞'] for i in live_list)
    
    video_hours = sum(i['时长(小时)'] for i in video_list)
    video_count = len(video_list)
    video_likes = sum(i['点赞'] for i in video_list)

    dates = pd.to_datetime(df_all['日期'])
    freq_str = "N/A"
    if len(dates) > 1:
        avg_days = round((dates.max() - dates.min()).days / len(dates), 1)
        freq_str = f"平均每 {avg_days} 天更新"

    # 构造 Rows
    summary_rows = [
        ["UP主昵称", nickname, ""],
        ["UID", uid, ""],
        ["统计年份", target_year or "全部历史", ""],
        ["粉丝数", fans, ""],
        ["总获赞数", total_likes, ""],
        ["总投币数", total_coins, ""],
        ["总收藏数", total_favs, ""],
        ["更新频率", freq_str, ""],
        
        ["--- 直播回放数据 ---", "", ""],
        ["回放合集名称", series_name, ""],
        ["直播场次", live_count, ""],
        ["直播总时长 (小时)", round(live_hours, 2), ""],
        ["场均时长 (小时)", round(live_hours/live_count, 2) if live_count else 0, ""],
        ["回放总获赞", live_likes, ""],
        
        ["--- 普通投稿数据 ---", "", ""],
        ["投稿数量", video_count, ""],
        ["投稿总时长 (小时)", round(video_hours, 2), ""],
        ["投稿总获赞", video_likes, ""]
    ]

    # 添加榜单
    rankings_config = [
        ("--- 播放榜 Top 5 (标题 | 播放量) ---", "播放"),
        ("--- 点赞榜 Top 5 (标题 | 点赞数) ---", "点赞"),
        ("--- 投币榜 Top 5 (标题 | 投币数) ---", "投币"),
        ("--- 收藏榜 Top 5 (标题 | 收藏数) ---", "收藏"),
        ("--- 三连榜 Top 5 (标题 | 综合分) ---", "三连"),
    ]

    for title, key in rankings_config:
        summary_rows.append([title, "", ""])
        top_list = get_top_n(all_data, key, 5)
        for item in top_list:
            summary_rows.append([item['标题'], item[key], item['链接']])
        if not top_list:
            summary_rows.append(["(无数据)", 0, ""])

    df_summary = pd.DataFrame(summary_rows, columns=["维度", "数值", "链接"])

    # 导出
    filename = f"{nickname}_{target_year or '全部'}_年度报告.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name="年度汇总", index=False)
        if live_list: pd.DataFrame(live_list).to_excel(writer, sheet_name="直播回放明细", index=False)
        if video_list: pd.DataFrame(video_list).to_excel(writer, sheet_name="普通投稿明细", index=False)
        
        title_text = str(target_year) if target_year else "历史全部数据"
        style_excel(writer, "年度汇总", title_text)

    print("-" * 40)
    print(f"🎉 报告生成成功: {filename}")

if __name__ == "__main__":
    if os.name == 'nt':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    try:
        sync(generate_report())
    except RuntimeError:
        loop = asyncio.get_event_loop()
        if loop.is_running(): asyncio.create_task(generate_report())
        else: loop.run_until_complete(generate_report())